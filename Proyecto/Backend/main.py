from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from fpdf import FPDF
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook
import datetime
import base64
import jwt
import time
import smtplib
import pymysql
import bcrypt
import os
import openpyxl
import io
import tempfile
import random

app = Flask(__name__)
CORS(app)

db_config = pymysql.connect(
    host='localhost',
    user='root',
    password='root',
    database='db_proyecto'
)

SECRET_KEY = os.urandom(24).hex()

def enviar_correo(destinatario, asunto, contenido):
    servidor_smtp = 'smtp.gmail.com'
    puerto = 587
    usuario = 'z07694496@gmail.com'
    password = 'kive hout kmgn pjpp'
    
    mensaje = MIMEMultipart()
    mensaje['From'] = usuario
    mensaje['To'] = destinatario
    mensaje['Subject'] = asunto
    mensaje.attach(MIMEText(contenido, 'plain'))
    
    try:
        with smtplib.SMTP(servidor_smtp, puerto) as server:
            server.starttls()
            server.login(usuario, password)
            server.sendmail(usuario, destinatario, mensaje.as_string())
        print("Correo enviado con éxito.")
    except Exception as e:
        print(f"Error al enviar el correo: {e}")

def rol_existe(rol_id):
    cursor = db_config.cursor()
    try:
        cursor.execute("SELECT COUNT(*) FROM rol WHERE id = %s", (rol_id,))
        return cursor.fetchone()[0] > 0
    except Exception as e:
        print(f"Error al verificar rol: {e}")
        return False
    finally:
        cursor.close()

def insertar_administrador():
    nombre = "Juan"
    apellido = "Pérez"
    dpi = "1234567890101"
    password = "contra12"  # Usa una contraseña hasheada válida
    rol_id = 1  # Asumiendo que el rol de administrador tiene id = 1

    # Encriptar la contraseña
    hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())

    cursor = db_config.cursor()
    try:
        # Verificar si ya existe el usuario
        cursor.execute("SELECT COUNT(*) FROM Usuario WHERE DPI = %s", (dpi,))
        if cursor.fetchone()[0] == 0:
            # Inserta el usuario administrador
            cursor.execute("""
                INSERT INTO Usuario (nombre, apellido, DPI, password, rol_id, login_attempts, account_locked)
                VALUES (%s, %s, %s, %s, %s, 0, FALSE)
            """, (nombre, apellido, dpi, hashed_password, rol_id))
            db_config.commit()
            print("Administrador insertado exitosamente.")
        else:
            print("El administrador ya existe.")
    except Exception as e:
        db_config.rollback()
        print(f"Error al insertar el administrador: {e}")
    finally:
        cursor.close()

@app.route('/recuperar_password', methods=['POST'])
def recuperar_password():
    # Obtiene los datos del cuerpo de la solicitud
    email = request.json.get('email')
    dpi = request.json.get('DPI')

    if not email and not dpi:
        return jsonify({"error": "Se requiere email o DPI."}), 400

    cursor = db_config.cursor()

    try:
        # Primero, intentar buscar por email en la tabla Estudiante
        if email:
            cursor.execute("SELECT * FROM Estudiante WHERE email = %s", (email,))
            estudiante = cursor.fetchone()
            if estudiante:
                # Generar token
                token = jwt.encode({
                    'user_id': estudiante[0],  # Asumiendo que el id es el primer campo
                    'exp': datetime.datetime.utcnow() + datetime.timedelta(hours=1)
                }, SECRET_KEY, algorithm='HS256')

                reset_link = f'http://localhost:8000/restablecer_password/{token}'
                enviar_correo(email, 'Restablecer contraseña', 
                              f'Haz clic en el siguiente enlace para restablecer tu contraseña: {reset_link}')

                return jsonify({"mensaje": "Si el email está registrado, se enviará un correo con instrucciones."}), 200

        # Si no se encuentra por email, buscar por DPI en la tabla Usuario
        if dpi:
            cursor.execute("SELECT * FROM Usuario WHERE DPI = %s", (dpi,))
            usuario = cursor.fetchone()
            if usuario:
                # Verificar si el usuario tiene un catedrático asociado
                cursor.execute("SELECT * FROM Catedratico WHERE usuario_id = %s", (usuario[0],))
                catedratico = cursor.fetchone()
                if catedratico:
                    # Generar token
                    token = jwt.encode({
                        'user_id': usuario[0],  # Asumiendo que el id es el primer campo
                        'exp': datetime.datetime.utcnow() + datetime.timedelta(hours=1)
                    }, SECRET_KEY, algorithm='HS256')

                    reset_link = f'http://localhost:8000/restablecer_password/{token}'
                    # Aquí puedes usar el email registrado en la tabla Usuario o un correo por defecto
                    enviar_correo('3508630320101@ingenieria.usac.edu.gt', 'Restablecer contraseña',  # usuario[4] sería el campo de email en la tabla Estudiante
                                  f'Haz clic en el siguiente enlace para restablecer tu contraseña: {reset_link}')

                    return jsonify({"mensaje": "Si el DPI está registrado, se enviará un correo con instrucciones."}), 200
                else:
                    return jsonify({"error": "El DPI proporcionado no corresponde a un catedrático."}), 404
        
        return jsonify({"error": "No se encontró ningún usuario asociado a los datos proporcionados."}), 404

    except Exception as e:
        return jsonify({"error": "Error al procesar la solicitud."}), 500

    finally:
        cursor.close()

@app.route('/actualizar_password', methods=['POST'])
def restablecer_password():
    token = request.json.get('token')
    nueva_password = request.json.get('nueva_password')

    if not nueva_password:
        return jsonify({"error": "La nueva contraseña es requerida."}), 400
    
    if len(nueva_password) < 8:
        return jsonify({"error": "La contraseña debe tener al menos 8 caracteres."}), 400

    try:
        data = jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
        user_id = data['user_id']

        hashed_password = bcrypt.hashpw(nueva_password.encode('utf-8'), bcrypt.gensalt())

        cursor = db_config.cursor()
        cursor.execute("UPDATE Usuario SET password = %s WHERE id = %s", (hashed_password, user_id))
        db_config.commit()

        return jsonify({"mensaje": "Contraseña restablecida exitosamente."}), 200

    except jwt.ExpiredSignatureError:
        return jsonify({"error": "El enlace ha expirado."}), 400
    except jwt.InvalidTokenError:
        return jsonify({"error": "Token inválido."}), 400
    except Exception as e:
        return jsonify({"error": f"Ocurrió un error inesperado: {str(e)}"}), 500

@app.route('/registro_estudiante', methods=['POST'])
def registrar_estudiante():
    datos = request.json
    # Obtener todos los datos necesarios del estudiante
    nombre = datos.get('nombre')
    apellido = datos.get('apellido')
    dpi = datos.get('DPI')
    fecha_nacimiento = datos.get('fecha_nacimiento')
    telefono = datos.get('telefono')
    nombre_usuario = datos.get('nombre_usuario')
    email = datos.get('email')
    password = datos.get('password')
    
    # Verifica si el nombre de usuario es único
    cursor = db_config.cursor()
    cursor.execute("SELECT COUNT(*) FROM Estudiante WHERE nombre_usuario = %s", (nombre_usuario,))
    if cursor.fetchone()[0] > 0:
        return jsonify({"error": "El nombre de usuario ya está en uso."}), 400

    # Encriptar la contraseña
    hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
    
    # Registro del usuario
    try:
        cursor.execute(""" 
            INSERT INTO Usuario (nombre, apellido, DPI, password, rol_id, login_attempts, account_locked)
            VALUES (%s, %s, %s, %s, %s, 0, FALSE)
        """, (nombre, apellido, dpi, hashed_password, 3))  # Asumimos rol_id = 3s para estudiantes
        usuario_id = cursor.lastrowid

        # Registro del estudiante
        cursor.execute(""" 
            INSERT INTO Estudiante (usuario_id, fecha_nacimiento, telefono, nombre_usuario, email)
            VALUES (%s, %s, %s, %s, %s)
        """, (usuario_id, fecha_nacimiento, telefono, nombre_usuario, email))
        
        db_config.commit()
        return jsonify({"mensaje": "Estudiante registrado exitosamente"}), 200
    except Exception as e:
        db_config.rollback()
        print(f"Error al insertar datos: {e}")
        return jsonify({"error": str(e)}), 400
    finally:
        cursor.close()
    
@app.route('/registro_catedratico', methods=['POST'])
def registrar_catedratico():
    datos = request.json
    print("Datos recibidos:", datos)  # Para depuración

    rol_id = datos.get('rol_id')
    print("Rol ID recibido:", rol_id)  # Para depuración

    # Cambia aquí: Permitir que el rol_id se envíe como 2
    if rol_id not in [1, 2]:  # Permitir registro solo si el rol es Administrador o Catedrático
        return jsonify({"error": "No tienes permiso para registrar un catedrático."}), 403

    nombre = datos.get('nombre')
    apellido = datos.get('apellido')
    dpi = datos.get('DPI')
    password = datos.get('password')
    especialidad = datos.get('especialidad')

    hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
    
    cursor = db_config.cursor()
    
    try:
        cursor.execute(""" 
            INSERT INTO Usuario (nombre, apellido, DPI, password, rol_id, login_attempts, account_locked)
            VALUES (%s, %s, %s, %s, %s, 0, FALSE)
        """, (nombre, apellido, dpi, hashed_password, 2))  # rol_id = 2 para catedráticos
        usuario_id = cursor.lastrowid

        print("Insertando catedrático...")
        cursor.execute(""" 
            INSERT INTO Catedratico (usuario_id, especialidad)
            VALUES (%s, %s)
        """, (usuario_id, especialidad))
        
        db_config.commit()
        return jsonify({"mensaje": "Catedrático registrado exitosamente"}), 200
    except Exception as e:
        db_config.rollback()
        print(f"Error al insertar datos: {e}")
        return jsonify({"error": str(e)}), 400
    finally:
        cursor.close()
    
@app.route('/login', methods=['POST'])
def iniciar_sesion():
    datos = request.json
    entrada = datos.get('nombre_usuario') or datos.get('DPI')  # Un solo campo para ambos
    password = datos.get('password').encode('utf-8')

    cursor = db_config.cursor(pymysql.cursors.DictCursor)

    # Buscar al usuario en la tabla Estudiante por nombre de usuario
    cursor.execute("SELECT * FROM Estudiante WHERE nombre_usuario = %s", (entrada,))
    estudiante = cursor.fetchone()

    # Si no se encuentra en Estudiante, buscar en Usuario por DPI
    if not estudiante:
        cursor.execute("SELECT * FROM Usuario WHERE DPI = %s", (entrada,))
        usuario = cursor.fetchone()
    else:
        # Si se encuentra, buscar los detalles de usuario asociado
        usuario_id = estudiante['usuario_id']  # Obtener el id del usuario asociado
        cursor.execute("SELECT * FROM Usuario WHERE id = %s", (usuario_id,))
        usuario = cursor.fetchone()

    if not usuario:
        return jsonify({"mensaje": "Usuario o contraseña incorrectos."}), 401
    
    print(f"Usuario encontrado: {usuario}")

    if usuario['account_locked']:
        return jsonify({"mensaje": "Cuenta bloqueada. Contacte al administrador."}), 403

    # Verificar la contraseña
    if bcrypt.checkpw(password, usuario['password'].encode('utf-8')):
        # Restablecer intentos de inicio de sesión
        cursor.execute("UPDATE Usuario SET login_attempts = 0 WHERE id = %s", (usuario['id'],))
        db_config.commit()
        
        # Obtener el rol del usuario
        rol_id = usuario.get('rol_id', 1)  # Asignar un valor por defecto si no se encuentra el rol
        nombre = usuario.get('nombre', 'Usuario desconocido')
        
        return jsonify({"mensaje": "Inicio de sesión exitoso", 
                        "id": usuario["id"],
                        "rol_id": rol_id,
                        "nombre": nombre}), 200
    else:
        # Incrementar intentos de inicio de sesión
        cursor.execute("UPDATE Usuario SET login_attempts = login_attempts + 1 WHERE id = %s", (usuario['id'],))
        db_config.commit()

        # Si los intentos fallidos son 3 o más, bloquear la cuenta
        if usuario['login_attempts'] >= 2:
            cursor.execute("UPDATE Usuario SET account_locked = TRUE WHERE id = %s", (usuario['id'],))
            db_config.commit()
            return jsonify({"mensaje": "Cuenta bloqueada. Contacte al administrador."}), 403

        return jsonify({"mensaje": "Usuario o contraseña incorrectos."}), 401

@app.route('/descargar_catedraticos', methods=['GET'])
def descargar_catedraticos():
    cursor = db_config.cursor()

    try:
        cursor.execute("SELECT u.nombre, u.apellido, u.DPI, c.especialidad FROM Usuario u JOIN Catedratico c ON u.id = c.usuario_id")
        catedraticos = cursor.fetchall()

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Catedraticos"
        sheet.append(["Nombre", "Apellido", "DPI", "Especialidad"])

        for catedratico in catedraticos:
            sheet.append(catedratico)

        file_stream = io.BytesIO()
        workbook.save(file_stream)
        file_stream.seek(0)

        return send_file(file_stream, as_attachment=True, download_name='catedraticos.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        print(f"Error al generar el archivo: {e}")
        return jsonify({"error": "Error al generar el archivo de catedráticos"}), 500
    finally:
        cursor.close()

@app.route('/listado_usuarios_bloqueados', methods=['GET'])
def listado_usuarios_bloqueados():
    cursor = db_config.cursor(pymysql.cursors.DictCursor)
    cursor.execute("SELECT id, nombre, apellido, DPI, rol_id FROM Usuario WHERE account_locked = TRUE AND rol_id != 1")  # Asumiendo rol_id = 1 es para administradores
    usuarios_bloqueados = cursor.fetchall()
    return jsonify(usuarios_bloqueados), 200

@app.route('/desbloquear/<int:usuario_id>', methods=['POST'])
def desbloquear_usuario(usuario_id):
    cursor = db_config.cursor()

    try:
        cursor.execute("SELECT * FROM Usuario WHERE id = %s", (usuario_id,))
        usuario = cursor.fetchone()

        if not usuario:
            return jsonify({"error": "Usuario no encontrado."}), 404
        
        # Verificar si el usuario es un administrador
        if usuario[5] == 1:  # Suponiendo que el rol_id del administrador es 1
            return jsonify({"error": "No se puede desbloquear a un administrador."}), 403
        
        cursor.execute("UPDATE Usuario SET account_locked = FALSE, login_attempts = 0 WHERE id = %s", (usuario_id,))
        db_config.commit()

        return jsonify({"mensaje": "Usuario desbloqueado exitosamente."}), 200

    except Exception as e:
        db_config.rollback()  # Hacer rollback en caso de error
        return jsonify({"error": str(e)}), 500

@app.route('/crear_curso', methods=['POST'])
def crear_curso():
    # Recibir los datos del curso desde la solicitud
    datos_curso = request.json
    
    # Extraer los campos necesarios
    nombre = datos_curso.get('nombre')
    codigo = datos_curso.get('codigo')
    costo = datos_curso.get('costo')
    horario = datos_curso.get('horario')
    cupo = datos_curso.get('cupo')
    catedratico_id = datos_curso.get('catedratico_id')
    banner_base64 = datos_curso.get('banner')
    mensaje_bienvenida = datos_curso.get('mensaje_bienvenida')
    estado = datos_curso.get('estado')

    # Imprimir los datos recibidos para verificar
    print("Datos del curso recibidos:")
    print(f"Nombre: {nombre}")
    print(f"Código: {codigo}")
    print(f"Costo: {costo}")
    print(f"Horario: {horario}")
    print(f"Cupo: {cupo}")
    print(f"Catedrático ID: {catedratico_id}")
    print(f"Mensaje de bienvenida: {mensaje_bienvenida}")
    print(f"Estado: {estado}")
    
    # Decodificar el banner en Base64
    if banner_base64:
        banner_data = base64.b64decode(banner_base64)
        # Guardar la imagen en un archivo
        banner_filename = f"banner_{codigo}.png"  # El nombre del archivo puede depender del código del curso, por ejemplo
        banner_path = os.path.join('C:/Users/Abdul Chacon/OneDrive - Facultad de Ingeniería de la Universidad de San Carlos de Guatemala/Documentos/Github/taller-python/Proyecto/banners', banner_filename)
        with open(banner_path, 'wb') as banner_file:
            banner_file.write(banner_data)
        print(f"Banner guardado en: {banner_path}")
    else:
        banner_path = None  # Si no se envía banner, puede ser opcional

    cursor = db_config.cursor()
    try:
        # Guardar el curso en la base de datos
        cursor.execute("""
            INSERT INTO Curso (nombre, codigo, costo, horario, cupo, catedratico_id, banner, mensaje_bienvenida, estado)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (nombre, codigo, costo, horario, cupo, catedratico_id, banner_path, mensaje_bienvenida, estado))
        db_config.commit()
        print("Curso insertado exitosamente.")
        return jsonify({"mensaje": "Curso creado exitosamente."}), 200
    except Exception as e:
        db_config.rollback()
        print(f"Error al insertar el curso: {e}")
        return jsonify({"error": "Error al crear el curso."}), 500
    finally:
        cursor.close()

@app.route('/listar_catedraticos_cursos', methods=['GET'])
def listar_catedraticos_cursos():
    cursor = db_config.cursor(pymysql.cursors.DictCursor)

    try:
        # Consultar la lista de catedráticos y los cursos que imparten
        cursor.execute("""
            SELECT 
                u.nombre AS catedratico_nombre,
                u.apellido AS catedratico_apellido,
                u.DPI AS catedratico_dpi,
                c.especialidad AS catedratico_especialidad,
                cu.nombre AS curso_nombre,
                cu.codigo AS curso_codigo,
                cu.horario AS curso_horario,
                cu.costo AS curso_costo,
                cu.cupo AS curso_cupo
            FROM Catedratico c
            JOIN Usuario u ON c.usuario_id = u.id
            LEFT JOIN Curso cu ON cu.catedratico_id = c.id
        """)
        resultados = cursor.fetchall()

        # Verificar si hay datos
        if not resultados:
            return jsonify({"mensaje": "No hay catedráticos o cursos registrados."}), 404

        # Devolver los resultados en formato JSON
        return jsonify(resultados), 200

    except Exception as e:
        print(f"Error al obtener la lista de catedráticos y cursos: {e}")
        return jsonify({"error": "Error al obtener la lista de catedráticos y cursos."}), 500

    finally:
        cursor.close()

@app.route('/profesor/cursos/<int:usuario_id>', methods=['GET'])
def obtener_cursos_profesor(usuario_id):
    cursor = db_config.cursor()
    try:
        cursor.execute("""
            SELECT c.id, c.nombre, c.codigo, c.costo, c.mensaje_bienvenida, c.horario, c.cupo, c.banner, c.estado, u.nombre
            FROM Curso c
            JOIN Catedratico ct ON c.catedratico_id = ct.id
            JOIN Usuario u ON ct.usuario_id = u.id
            WHERE ct.usuario_id = %s
        """, (usuario_id,))
        
        cursos = cursor.fetchall()

        if not cursos:  # Si no se encuentran cursos
            print(f"No se encontraron cursos para el usuario ID: {usuario_id}")  # Para depuración
            return jsonify({"message": "No se encontraron cursos para este catedrático."}), 404

        lista_cursos = []
        for curso in cursos:
            lista_cursos.append({
                'id': curso[0],
                'nombre': curso[1],
                'codigo': curso[2],
                'costo': curso[3],
                'descripcion': curso[4],  # Ahora usa `mensaje_bienvenida`
                'horario': curso[5],
                'cupo': curso[6],
                'banner': curso[7],  # El banner será una ruta a la imagen
                'estado': curso[8],
                'catedratico': curso[9]
            })

        return jsonify(lista_cursos), 200

    except Exception as e:
        print(f"Error al obtener cursos: {str(e)}")  # Imprimir el error en la consola
        return jsonify({"error": f"Ocurrió un error inesperado: {str(e)}"}), 500
    finally:
        cursor.close()

@app.route('/profesor/editar_banner', methods=['POST'])
def editar_banner():
    datos = request.json
    curso_id = datos.get('curso_id')
    nuevo_banner = datos.get('banner')

    if not curso_id or not nuevo_banner:
        return jsonify({"error": "El curso_id y el banner son requeridos."}), 400

    cursor = db_config.cursor()
    try:
        # Actualizar el banner del curso
        cursor.execute("UPDATE Curso SET banner = %s WHERE id = %s", (nuevo_banner, curso_id))
        db_config.commit()

        return jsonify({"mensaje": "Banner actualizado exitosamente."}), 200

    except Exception as e:
        db_config.rollback()
        print(f"Error al actualizar el banner: {e}")
        return jsonify({"error": "Error al actualizar el banner."}), 500

    finally:
        cursor.close()

@app.route('/cursos', methods=['GET'])
def listar_cursos():
    cursor = db_config.cursor(pymysql.cursors.DictCursor)
    try:
        cursor.execute("SELECT id, nombre, codigo, costo, cupo FROM Curso WHERE estado = 1")  # Asegúrate de incluir todos los campos necesarios
        cursos = cursor.fetchall()
        return jsonify(cursos), 200
    except Exception as e:
        print(f"Error al obtener la lista de cursos: {e}")
        return jsonify({"error": "Error al obtener la lista de cursos"}), 500
    finally:
        cursor.close()
        
@app.route('/estudiante/inscribir', methods=['POST'])
def inscribir_curso():
    datos = request.json
    estudiante_id = datos.get('estudiante_id')
    curso_id = datos.get('curso_id')
    
    if not estudiante_id or not curso_id:
        return jsonify({"error": "Se requiere estudiante_id y curso_id."}), 400

    cursor = db_config.cursor()
    try:
        # Verificar que el curso existe y tiene cupo disponible
        cursor.execute("SELECT cupo FROM Curso WHERE id = %s AND estado = 1", (curso_id,))
        curso = cursor.fetchone()
        if not curso:
            return jsonify({"error": "El curso no existe o no está disponible."}), 404
        
        if curso[0] <= 0:
            return jsonify({"error": "El curso no tiene cupos disponibles."}), 400

        # Verificar si el estudiante ya está inscrito en el curso
        cursor.execute("SELECT COUNT(*) FROM Inscripcion WHERE usuario_id = %s AND curso_id = %s", (estudiante_id, curso_id))
        if cursor.fetchone()[0] > 0:
            return jsonify({"error": "Ya estás inscrito en este curso."}), 400

        # Obtener la fecha de inscripción usando time
        fecha_inscripcion = time.strftime('%Y-%m-%d')  # Formato de fecha YYYY-MM-DD

        # Registrar la inscripción
        cursor.execute("INSERT INTO Inscripcion (usuario_id, curso_id, fecha_inscripcion) VALUES (%s, %s, %s)", 
                       (estudiante_id, curso_id, fecha_inscripcion))
        # Actualizar el cupo del curso
        cursor.execute("UPDATE Curso SET cupo = cupo - 1 WHERE id = %s", (curso_id,))
        db_config.commit()

        # Enviar correo de confirmación
        # Obtener el email del estudiante
        cursor.execute("SELECT email FROM Estudiante WHERE usuario_id = %s", (estudiante_id,))
        estudiante = cursor.fetchone()
        if estudiante and estudiante[0]:
            enviar_correo(
                estudiante[0],
                'Confirmación de inscripción',
                f'Te has inscrito exitosamente en el curso con ID {curso_id}.'
            )
        
        return jsonify({"mensaje": "Inscripción exitosa."}), 200
    except Exception as e:
        db_config.rollback()
        print(f"Error al inscribirse en el curso: {e}")
        return jsonify({"error": "Error al inscribirse en el curso."}), 500
    finally:
        cursor.close()

@app.route('/estudiante/cursos/<int:estudiante_id>', methods=['GET'])
def cursos_inscritos(estudiante_id):
    cursor = db_config.cursor(pymysql.cursors.DictCursor)
    try:
        cursor.execute("""
            SELECT c.* FROM Curso c
            JOIN Inscripcion i ON c.id = i.curso_id
            WHERE i.usuario_id = %s
        """, (estudiante_id,))
        cursos = cursor.fetchall()
        return jsonify(cursos), 200
    except Exception as e:
        print(f"Error al obtener los cursos inscritos: {e}")
        return jsonify({"error": "Error al obtener los cursos inscritos."}), 500
    finally:
        cursor.close()

@app.route('/estudiante/desinscribir', methods=['POST'])
def desinscribir_curso():
    datos = request.json
    estudiante_id = datos.get('estudiante_id')
    curso_id = datos.get('curso_id')
    
    # Verificar que estudiante_id y curso_id estén presentes
    if not estudiante_id or not curso_id:
        return jsonify({"error": "Se requiere estudiante_id y curso_id."}), 400

    cursor = db_config.cursor()
    try:
        # Verificar que el estudiante está inscrito en el curso
        cursor.execute("SELECT COUNT(*) FROM Inscripcion WHERE usuario_id = %s AND curso_id = %s", (estudiante_id, curso_id))
        if cursor.fetchone()[0] == 0:
            return jsonify({"error": "No estás inscrito en este curso."}), 400

        # Eliminar la inscripción
        cursor.execute("DELETE FROM Inscripcion WHERE usuario_id = %s AND curso_id = %s", (estudiante_id, curso_id))
        
        # Incrementar el cupo del curso
        cursor.execute("UPDATE Curso SET cupo = cupo + 1 WHERE id = %s", (curso_id,))
        
        db_config.commit()
        return jsonify({"mensaje": "Desinscripción exitosa."}), 200

    except Exception as e:
        db_config.rollback()
        print(f"Error al desinscribirse del curso: {e}")
        return jsonify({"error": "Error al desinscribirse del curso."}), 500
    finally:
        cursor.close()

@app.route('/estudiante/certificado/<int:curso_id>/<int:estudiante_id>', methods=['GET'])
def descargar_certificado(curso_id, estudiante_id):
    cursor = db_config.cursor()
    try:
        # Obtener la nota del estudiante en el curso
        cursor.execute("""
            SELECT n.nota FROM Nota n
            JOIN Inscripcion i ON n.inscripcion_id = i.id
            WHERE i.curso_id = %s AND i.usuario_id = %s
        """, (curso_id, estudiante_id))
        nota_result = cursor.fetchone()
        if not nota_result:
            return jsonify({"error": "No se encontró la nota del estudiante en este curso."}), 404

        nota = nota_result[0]
        if nota < 61:
            return jsonify({"error": "La nota es insuficiente para obtener el certificado."}), 400

        # Obtener datos del estudiante y del curso
        cursor.execute("SELECT nombre, apellido FROM Usuario WHERE id = %s", (estudiante_id,))
        estudiante = cursor.fetchone()
        cursor.execute("SELECT nombre FROM Curso WHERE id = %s", (curso_id,))
        curso = cursor.fetchone()

        # Generar el certificado (ejemplo simple con PDF)
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=16)
        pdf.cell(200, 10, txt="Certificado de Aprobación", ln=1, align='C')
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt=f"Estudiante: {estudiante[0]} {estudiante[1]}", ln=2, align='L')
        pdf.cell(200, 10, txt=f"Curso: {curso[0]}", ln=3, align='L')
        pdf.cell(200, 10, txt=f"Nota: {nota}", ln=4, align='L')

        # Crear un archivo temporal para guardar el PDF
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
            pdf.output(tmp_file.name)  # Guardar el PDF en el archivo temporal
            tmp_file.seek(0)  # Mover el puntero al inicio del archivo

            # Enviar el PDF generado
            return send_file(
                tmp_file.name,
                as_attachment=True,
                download_name='certificado.pdf',
                mimetype='application/pdf'
            )
    except Exception as e:
        print(f"Error al generar el certificado: {e}")
        return jsonify({"error": "Error al generar el certificado."}), 500
    finally:
        cursor.close()

@app.route('/estudiante/completar_curso/<int:curso_id>', methods=['POST'])
def completar_curso(curso_id):
    # Obtener el usuario_id desde el cuerpo de la solicitud JSON
    data = request.get_json()
    usuario_id = data.get('usuario_id')

    if not usuario_id:
        return jsonify({"error": "Debes proporcionar tu ID de usuario."}), 403

    # Generar una nota aleatoria
    nota = random.uniform(0, 100)
    nota = round(nota, 2)  # Redondear a 2 decimales

    # Obtener el inscripcion_id basado en usuario_id y curso_id
    inscripcion_id = obtener_inscripcion_id(usuario_id, curso_id)
    if not inscripcion_id:
        return jsonify({"error": "Inscripción no encontrada para este curso."}), 404

    # Insertar la nota en la base de datos
    cursor = db_config.cursor()
    try:
        cursor.execute("INSERT INTO Nota (inscripcion_id, nota) VALUES (%s, %s)", (inscripcion_id, nota))
        db_config.commit()

        # Verificar si la nota es suficiente para generar el certificado
        if nota >= 61:
            # Generar certificado
            cursor.execute("INSERT INTO Certificado (inscripcion_id, fecha_emision) VALUES (%s, NOW())", (inscripcion_id,))
            db_config.commit()

        return jsonify({"nota": nota, "mensaje": "Curso completado. ¡Bien hecho!"})

    except Exception as e:
        db_config.rollback()  # Deshacer cambios en caso de error
        return jsonify({"error": str(e)}), 500

    finally:
        cursor.close()

def obtener_inscripcion_id(usuario_id, curso_id):
    # Implementa esta función para obtener el inscripcion_id basado en usuario_id y curso_id
    cursor = db_config.cursor()
    cursor.execute("SELECT id FROM Inscripcion WHERE usuario_id = %s AND curso_id = %s", (usuario_id, curso_id))
    result = cursor.fetchone()
    cursor.close()
    
    return result[0] if result else None

@app.route('/profesor/descargar_registro_notas/<int:curso_id>', methods=['GET'])
def descargar_registro_notas(curso_id):
    """Descarga el registro de notas de un curso en formato .xlsx."""
    cursor = db_config.cursor()

    try:
        # Verificar si el curso existe
        cursor.execute("SELECT COUNT(*) FROM Curso WHERE id = %s", (curso_id,))
        if cursor.fetchone()[0] == 0:
            return jsonify({"error": "Curso no encontrado."}), 404
        
        # Obtener las notas de los estudiantes en el curso
        cursor.execute("""
            SELECT u.nombre, u.apellido, n.nota
            FROM Usuario u
            JOIN Inscripcion i ON u.id = i.usuario_id
            JOIN Nota n ON i.id = n.inscripcion_id
            WHERE i.curso_id = %s
        """, (curso_id,))
        resultados = cursor.fetchall()

        # Manejar caso de no encontrar resultados
        if not resultados:
            return jsonify({"error": "No hay notas disponibles para este curso."}), 404

        # Crear un libro de trabajo y una hoja
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registro de Notas"

        # Agregar encabezados
        ws.append(["Nombre", "Apellido", "Nota"])

        # Agregar los datos
        for nombre, apellido, nota in resultados:
            ws.append([nombre, apellido, nota])

        # Crear un objeto BytesIO para almacenar el archivo en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Devolver el archivo Excel como respuesta
        return send_file(
            output,
            as_attachment=True,
            download_name=f'registro_notas_curso_{curso_id}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        print(f"Error al generar el registro de notas: {e}")
        return jsonify({"error": "Error al generar el registro de notas."}), 500
    finally:
        cursor.close()
        
@app.route('/notas/<int:curso_id>', methods=['GET'])
def obtener_notas(curso_id):
    cursor = db_config.cursor(pymysql.cursors.DictCursor)
    try:
        cursor.execute("""
            SELECT Usuario.nombre, Usuario.apellido, Nota.nota AS calificacion
            FROM Nota
            JOIN Inscripcion ON Nota.inscripcion_id = Inscripcion.id
            JOIN Estudiante ON Inscripcion.usuario_id = Estudiante.usuario_id
            JOIN Usuario ON Estudiante.usuario_id = Usuario.id
            WHERE Inscripcion.curso_id = %s
        """, (curso_id,))
        notas = cursor.fetchall()
        return jsonify(notas), 200
    except Exception as e:
        print(f"Error al obtener las notas del curso {curso_id}: {e}")
        return jsonify({"error": "Error al obtener las notas"}), 500
    finally:
        cursor.close()

@app.route('/descargar_notas/<int:curso_id>', methods=['GET'])
def descargar_notas(curso_id):
    cursor = db_config.cursor(pymysql.cursors.DictCursor)
    try:
        cursor.execute("""
            SELECT Usuario.nombre, Usuario.apellido, Nota.nota AS calificacion
            FROM Nota
            JOIN Inscripcion ON Nota.inscripcion_id = Inscripcion.id
            JOIN Estudiante ON Inscripcion.usuario_id = Estudiante.usuario_id
            JOIN Usuario ON Estudiante.usuario_id = Usuario.id
            WHERE Inscripcion.curso_id = %s
        """, (curso_id,))
        notas = cursor.fetchall()

        # Crear un libro de trabajo de Excel
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Notas"

        # Encabezados
        sheet.append(["Nombre", "Apellido", "Calificación"])

        # Agregar los datos
        for nota in notas:
            sheet.append([nota["nombre"], nota["apellido"], nota["calificacion"]])

        # Guardar el archivo en memoria
        archivo_excel = f"notas_curso_{curso_id}.xlsx"
        workbook.save(archivo_excel)

        # Devolver el archivo como una respuesta
        return send_file(archivo_excel, as_attachment=True)

    except Exception as e:
        print(f"Error al descargar las notas del curso {curso_id}: {e}")
        return jsonify({"error": "Error al descargar las notas"}), 500
    finally:
        cursor.close()

if __name__ == '__main__':
    insertar_administrador()
    app.run(host='0.0.0.0', port=3000, debug=True)